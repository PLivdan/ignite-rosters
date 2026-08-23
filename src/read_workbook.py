"""Read out/ignite-rosters.xlsx back into build/current.json.

This is what makes an edit round-trip: change a name in the workbook, re-run,
and every card carrying that name is redrawn.
"""
import json, os, sys
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from openpyxl import load_workbook
from common import BUILD, OUT, RAW

ROLE_ORDER = {"DPS": 0, "Tank": 1, "Support": 2, "Flex": 3,
              "Head Coach": 4, "Coach": 5, "Assistant Coach": 6, "Analyst": 7}


def main():
    path = os.path.join(OUT, "ignite-rosters.xlsx")
    if not os.path.exists(path):
        raise SystemExit(f"no workbook at {path} — run build_workbook.py first")
    cur = json.load(open(os.path.join(BUILD, "current.json")))
    wb = load_workbook(path, data_only=True)
    ws = wb["Rosters"]

    icons = {}
    for t in cur["teams"]:
        for p in t["roster"]:
            if p.get("icon"):
                icons[p["name"].lower()] = p["icon"]

    rosters, seen_teams = {}, []
    for row in ws.iter_rows(min_row=2, values_only=True):
        region, team, short, role, player, status = (list(row) + [None] * 6)[:6]
        confirmed, since, source, notes = (list(row) + [None] * 10)[6:10]
        if not team:
            continue
        if team not in rosters:
            rosters[team] = []
            seen_teams.append(team)
        if not player or not str(player).strip():
            continue          # an unfilled slot stays unfilled
        st = (status or "main").strip()
        rosters[team].append({
            "name": str(player).strip(),
            "role": (role or "").strip(),
            "status": "main" if st == "LEAKED" else st,
            "status_flag": "leaked" if st == "LEAKED" else "",
            "origin": "workbook",
            "since": str(since or ""),
            "ref": "",
            "icon": icons.get(str(player).strip().lower(), ""),
        })

    changed = 0
    for t in cur["teams"]:
        new = rosters.get(t["name"])
        if new is None:
            continue
        old_names = [p["name"] for p in t["roster"]]
        new.sort(key=lambda p: ({"main": 0, "sub": 1, "staff": 2}.get(p["status"], 3),
                                ROLE_ORDER.get(p["role"], 9), p["name"].lower()))
        if [p["name"] for p in new] != old_names:
            changed += 1
        t["roster"] = new
        mains = [p for p in new if p["status"] == "main"]
        by_role = {}
        for p in mains:
            by_role[p["role"]] = by_role.get(p["role"], 0) + 1
        t["mains"] = len(mains)
        t["needs"] = max(0, 6 - len(mains))
        t["contested"] = [f"{n}x {r}" for r, n in by_role.items() if n > 2]
        t["complete"] = len(mains) == 6 and not t["contested"]

    cur["source_of_truth"] = "workbook"
    with open(os.path.join(BUILD, "current.json"), "w") as f:
        json.dump(cur, f, indent=2, ensure_ascii=False)
    print(f"read {sum(len(v) for v in rosters.values())} rows across "
          f"{len(seen_teams)} teams; {changed} rosters changed")


if __name__ == "__main__":
    main()
