"""out/ignite-rosters.xlsx — the editable source of truth for the Stage 2 card set."""
import json, os, sys
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from common import BUILD, OUT
import leaks as leaks_mod

HEAD_FILL = PatternFill("solid", fgColor="1E2530")
HEAD_FONT = Font(color="FFFFFF", bold=True, size=11)
LEAK_FILL = PatternFill("solid", fgColor="FFF2D9")
LEAK_FONT = Font(color="8A5A00", bold=True)
GAP_FILL = PatternFill("solid", fgColor="FDECEA")
NEW_FILL = PatternFill("solid", fgColor="E8F3EC")
STAFF_FONT = Font(color="6B7280", italic=True)
THIN = Border(bottom=Side("thin", color="E3E6EA"))


def style(ws, widths, wrap_cols=()):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for c in ws[1]:
        c.fill, c.font = HEAD_FILL, HEAD_FONT
        c.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 22
    for row in ws.iter_rows(min_row=2):
        for i in wrap_cols:
            row[i].alignment = Alignment(wrap_text=True, vertical="top")


def main():
    os.makedirs(OUT, exist_ok=True)
    data = json.load(open(os.path.join(BUILD, "current.json")))
    resolved = leaks_mod.resolve(data, leaks_mod.load())
    teams = data["teams"]
    region_of = {t["name"]: t["region"] for t in teams}

    wb = Workbook()

    # ---- Rosters -------------------------------------------------------
    ws = wb.active
    ws.title = "Rosters"
    ws.append(["Region", "Team", "Short", "Role", "Player", "Status",
               "Confirmed", "Since", "Source", "Notes"])
    for t in teams:
        for p in t["roster"]:
            src = ("Liquipedia Stage 1" if p["origin"] == "stage1"
                   else f"Transfer list {p['since']}")
            ws.append([t["region"], t["name"], t["short"], p["role"], p["name"],
                       p["status"], "yes", p.get("since", ""), src, ""])
        # explicit blank slots for whatever the team is still missing
        for _ in range(t["needs"]):
            ws.append([t["region"], t["name"], t["short"], "", "", "main", "no", "",
                       "UNFILLED", "type the name here, then re-run build_all.py"])
    # A leaked name is only worth a row if the player is not already on that
    # team; otherwise the leak duplicates a confirmed starter and inflates the roster.
    on_team = {t["name"]: {p["name"].lower() for p in t["roster"]} for t in teams}
    for r in resolved:
        if r["kind"] != "roster":
            continue
        for nm in r["players"]:
            if nm.lower() in on_team.get(r["team"], set()):
                continue
            ws.append([region_of.get(r["team"], ""), r["team"], "", "", nm, "LEAKED",
                       "no", r["date"], f"AkamaruRivals {r['date']}",
                       f"{r['claim']} — {r['verdict']}"])
    for row in ws.iter_rows(min_row=2):
        row[0].border = row[4].border = THIN
        if row[5].value == "LEAKED":
            for c in row:
                c.fill, c.font = LEAK_FILL, LEAK_FONT
        elif row[8].value == "UNFILLED":
            for c in row:
                c.fill = GAP_FILL
        elif row[5].value == "staff":
            for c in row:
                c.font = STAFF_FONT
        elif str(row[8].value).startswith("Transfer"):
            row[8].font = Font(color="1B7F4B")
    dv = DataValidation(type="list", formula1='"main,sub,staff,LEAKED"', allow_blank=True)
    ws.add_data_validation(dv); dv.add(f"F2:F{ws.max_row}")
    dv2 = DataValidation(type="list", formula1='"yes,no"', allow_blank=True)
    ws.add_data_validation(dv2); dv2.add(f"G2:G{ws.max_row}")
    dv3 = DataValidation(type="list",
                         formula1='"DPS,Tank,Support,Flex,Head Coach,Coach,Assistant Coach,Analyst"',
                         allow_blank=True)
    ws.add_data_validation(dv3); dv3.add(f"D2:D{ws.max_row}")
    style(ws, [9, 24, 8, 15, 20, 10, 11, 12, 22, 44], wrap_cols=(9,))

    # ---- Teams ---------------------------------------------------------
    ws = wb.create_sheet("Teams")
    ws.append(["Region", "Team", "Short", "Starters", "Still needed", "Contested",
               "Subs", "Staff", "Roster status", "Twitter", "Twitch", "Website"])
    for t in teams:
        c = {s: sum(1 for p in t["roster"] if p["status"] == s)
             for s in ("main", "sub", "staff")}
        st = ("complete" if t["complete"]
              else "over-full" if t["contested"]
              else "no roster published" if c["main"] == 0
              else f"missing {t['needs']}")
        L = t.get("socials") or {}
        ws.append([t["region"], t["name"], t["short"], c["main"], t["needs"],
                   ", ".join(t["contested"]), c["sub"], c["staff"], st,
                   L.get("twitter", ""), L.get("twitch", ""), L.get("home", "")])
    for row in ws.iter_rows(min_row=2):
        v = str(row[8].value)
        if v == "complete":
            row[8].font = Font(color="1B7F4B", bold=True)
        else:
            for c in row:
                c.fill = GAP_FILL
            row[8].font = Font(color="B3261E", bold=True)
    style(ws, [9, 24, 8, 10, 13, 14, 7, 7, 22, 30, 30, 28])

    # ---- Power ranking -------------------------------------------------
    ws = wb.create_sheet("Power ranking")
    ws.append(["Rank", "Region", "Team", "Tier", "Roster status", "Verdict", "Notes"])
    for region in ("EU", "NA"):
        for t in [x for x in teams if x["region"] == region]:
            ws.append(["", region, t["name"], "",
                       "complete" if t["complete"] else f"missing {t['needs']}", "", ""])
    dv4 = DataValidation(type="list", formula1='"S,A,B,C,D,F"', allow_blank=True)
    ws.add_data_validation(dv4); dv4.add(f"D2:D{ws.max_row}")
    style(ws, [7, 9, 24, 8, 16, 34, 54])

    # ---- Roster changes -------------------------------------------------
    ws = wb.create_sheet("Roster changes")
    ws.append(["Date", "Player", "What happened"])
    for date, name, what in sorted(data["transfer_effects"], reverse=True):
        ws.append([date, name, what])
    for row in ws.iter_rows(min_row=2):
        if str(row[2].value).startswith("joined"):
            row[2].font = Font(color="1B7F4B")
        else:
            row[2].font = Font(color="B3261E")
    style(ws, [13, 22, 46])

    # ---- Leaks ----------------------------------------------------------
    ws = wb.create_sheet("Leaks")
    ws.append(["Date", "Team", "Players", "Claim", "Verdict", "Landed",
               "What happened", "Post", "Caveat"])
    for r in resolved:
        ws.append([r["date"], r.get("team", ""), ", ".join(r.get("players", [])),
                   r.get("claim", ""), r["verdict"],
                   f"{r.get('landed','')}/{r.get('claimed','')}" if r["kind"] == "roster" else "",
                   r["detail"], r["text"], r.get("caveat", "")])
    for row in ws.iter_rows(min_row=2):
        v = row[4].value
        col = {"landed": "1B7F4B", "partial": "B26A00",
               "not reflected": "B3261E"}.get(v)
        if col:
            row[4].font = Font(color=col, bold=True)
    style(ws, [12, 20, 28, 14, 15, 9, 54, 54, 44], wrap_cols=(6, 7, 8))

    # ---- Read me --------------------------------------------------------
    ws = wb.create_sheet("Read me")
    for r in [
        ["Marvel Rivals Ignite 2026 Stage 2, EU and NA"],
        [f"Built {data['generated']}"],
        [""],
        ["HOW TO USE"],
        ["  1. Edit the Rosters sheet — type into the pink UNFILLED rows, or change any name."],
        ["  2. Run:  python3 src/build_all.py --from-workbook"],
        ["  3. out/cards/ is rebuilt from this file. Re-zip and send."],
        [""],
        ["WHY ROWS ARE MISSING"],
        ["  Liquipedia has published the Stage 2 team list but not the Stage 2 rosters."],
        ["  Names here are the Stage 1 roster with every 2026 Q3 transfer applied on top —"],
        ["  the best available picture, not an official Stage 2 roster. Pink rows are"],
        ["  genuinely unknown and are where you type live during the video."],
        [""],
        ["STATUS VALUES"],
        ["  main    starter (a full team is 2 Duelist, 2 Vanguard, 2 Strategist)"],
        ["  sub     substitute"],
        ["  staff   coach / assistant / analyst"],
        ["  LEAKED  reported, unconfirmed — renders amber on the cards"],
        [""],
        ["SOURCES"],
        ["  Team list       Liquipedia MR Ignite/2026/Stage 2 (EMEA, Americas)"],
        ["  Roster baseline Liquipedia MR Ignite/2026/Stage 1"],
        ["  Corrections     Liquipedia Player Transfers/2026/3rd Quarter"],
        ["  Logos           Liquipedia (full resolution, light + dark variants)"],
                ["  Leaks           x.com/AkamaruRivals, pasted by hand"],
        [""],
        ["KNOWN GAPS"],
        ["  No logo on any source, so cards use a lettered monogram:"],
        ["     The Chosen Ones, JollyJesters"],
        ["  Over-full rosters (source spells a departure differently than the arrival):"],
        ["     CrownFall and Vengeful each show 3 Supports. Delete the wrong one."],
    ]:
        ws.append(r)
    ws["A1"].font = Font(bold=True, size=14)
    for r in (4, 9, 15, 21, 29):
        ws.cell(row=r, column=1).font = Font(bold=True)
    ws.column_dimensions["A"].width = 96

    path = os.path.join(OUT, "ignite-rosters.xlsx")
    wb.save(path)
    print(f"wrote {path}")
    for s in wb.sheetnames:
        print(f"   {s:<18} {wb[s].max_row - 1} rows")


if __name__ == "__main__":
    main()
